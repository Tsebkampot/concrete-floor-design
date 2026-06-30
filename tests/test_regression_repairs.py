"""只读校核报告对应的回归测试。"""

from copy import deepcopy
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
import zipfile

import pandas as pd
import pytest
from openpyxl import load_workbook

import app
from calculations.common import default_parameters
from calculations.force_envelope import analyze_member_envelope
from calculations.matrix_stiffness import BeamElement, BeamNode, solve_continuous_beam
from calculations.project_analysis import analyze_project_matrix, parse_supports, slab_support_line_reactions
from calculations.rebar import recommend_longitudinal_rebar
from calculations.section_design import design_control_sections
from calculations.slab import SlabInput, calculate_loads
from calculations.structural_models import ContinuousMemberModel, MemberPointLoad, analyze_member
from charts.plot_control_sections import plot_control_section_diagram
from export.export_excel import build_excel_workbook
from export.export_pdf import build_pdf_report
from export.export_report import build_markdown_report
from export.report_summary import REPORT_SUMMARY_COLUMNS, build_calculation_book_tables
from export.export_word import build_word_report


EI = 100_000.0


def _params_with(**updates):
    params = deepcopy(default_parameters().to_dict())
    params.update(updates)
    return params


def _large_21x36_params(**updates):
    params = _params_with(
        l1_m=21.0,
        l2_m=36.0,
        slab_span_m=3.5,
        secondary_span_m=9.0,
        main_span_m=7.0,
        slab_spans_text="3,4",
        slab_supports_text="pin,pin,pin",
        slab_live_spans_text="1,2",
        secondary_spans_text="9,9,9,9",
        secondary_supports_text="pin,pin,pin,pin,pin",
        secondary_live_spans_text="1,2,3,4",
        main_spans_text="7,7,7",
        main_supports_text="pin,pin,pin,pin",
        main_live_spans_text="1,2,3",
    )
    params.update(updates)
    return params


def _manual_20x28_params(**updates):
    params = _params_with(
        l1_m=20.0,
        l2_m=28.0,
        slab_span_m=2.5,
        secondary_span_m=7.0,
        main_span_m=5.0,
        slab_spans_text="2.5,2.5",
        slab_supports_text="pin,pin,pin",
        slab_live_spans_text="1",
        secondary_spans_text="7,7,7,7",
        secondary_supports_text="pin,pin,pin,pin,pin",
        secondary_live_spans_text="2,4",
        main_spans_text="5,5,5,5",
        main_supports_text="pin,pin,pin,pin,pin",
        main_live_spans_text="1,3",
        automatic_live_patterns=False,
    )
    params.update(updates)
    return params


def _slab_model(strip_width: float, spans=(2.0, 2.0, 2.0)) -> ContinuousMemberModel:
    return ContinuousMemberModel(
        "板", spans, EI * strip_width,
        tuple(3.0 * strip_width for _ in spans), tuple(4.0 * strip_width for _ in spans),
        tuple(0.2 for _ in range(len(spans) + 1)), tuple("pin" for _ in range(len(spans) + 1)),
    )


@pytest.mark.parametrize("strip_width", [0.5, 1.0, 2.0])
def test_slab_transfer_reaction_is_independent_of_strip_width(strip_width: float) -> None:
    model = _slab_model(strip_width)
    case = analyze_member(model, set(range(3)), include_dead=True)
    values = slab_support_line_reactions(model, case, strip_width)
    reference_model = _slab_model(1.0)
    reference = slab_support_line_reactions(reference_model, analyze_member(reference_model, set(range(3)), True), 1.0)
    assert values == pytest.approx(reference)


def test_unequal_slab_support_reactions_are_not_replaced_by_one_maximum() -> None:
    model = _slab_model(1.0, (1.5, 2.0, 2.5))
    reactions = slab_support_line_reactions(model, analyze_member(model, set(range(3)), True), 1.0)
    assert len({round(value, 5) for value in reactions}) > 1


def test_global_case_name_is_continuous_through_three_levels() -> None:
    project = analyze_project_matrix(default_parameters().to_dict(), calculate_loads(SlabInput()))
    names = set(project.global_case_df["全局工况"])
    assert {"G", "Q", "G+Q[1,2,3]"}.issubset(names)
    assert names == set(project.slab.cases) == set(project.secondary.cases) == set(project.main.cases)
    assert set(project.transfer_df["全局工况"]) == names
    assert project.transfer_df["荷载来源"].str.contains("板支承线").all()


def test_support_center_and_true_faces_are_control_sections() -> None:
    model = ContinuousMemberModel("次梁", (5.0, 5.0), EI, (8.0, 8.0), (4.0, 4.0), (0.3, 0.4, 0.3), ("pin",) * 3)
    envelope = analyze_member_envelope(model)
    centers = [s for s in envelope.control_sections if s.name == "内支座中心截面"]
    assert centers and centers[0].x_m == pytest.approx(5.0)
    left_face = next(s for s in envelope.control_sections if s.name == "右支座左边缘截面" and s.span_index == 0)
    right_face = next(s for s in envelope.control_sections if s.name == "左支座右边缘截面" and s.span_index == 1)
    assert left_face.x_m == pytest.approx(5.0 - 0.4 / 2)
    assert right_face.x_m == pytest.approx(5.0 + 0.4 / 2)
    center_row = envelope.control_df[envelope.control_df["截面名称"] == "内支座中心截面"]
    assert not center_row.empty and (center_row["最大负弯矩 (kN·m)"] < 0).any()


def test_overlapping_support_faces_raise_instead_of_clipping_to_point_two_l() -> None:
    with pytest.raises(ValueError, match="净跨|重叠"):
        ContinuousMemberModel("次梁", (0.5,), EI, (1.0,), (0.0,), (0.5, 0.5), ("pin", "pin"))


def test_secondary_and_main_control_diagrams_use_distinct_directions() -> None:
    project = analyze_project_matrix(default_parameters().to_dict(), calculate_loads(SlabInput()))
    secondary_chart = plot_control_section_diagram(project.secondary)
    main_chart = plot_control_section_diagram(project.main)
    assert "次梁控制截面示意图（30m方向）" in secondary_chart.svg
    assert "主梁控制截面示意图（18m方向）" in main_chart.svg
    assert project.secondary.model.boundaries_m[-1] == pytest.approx(30.0)
    assert project.main.model.boundaries_m[-1] == pytest.approx(18.0)
    assert all(mark in secondary_chart.svg for mark in (">1<", ">B<", ">2<", ">C<", ">3<"))


def test_support_count_mismatch_is_explicit_error() -> None:
    with pytest.raises(ValueError, match="数量应为 4"):
        parse_supports("fixed,pin", 3, "次梁支承条件")


def test_failed_input_invalidates_current_result_and_blocks_export() -> None:
    state = SimpleNamespace(current_results={"old": True}, current_tables={"old": True}, latest_results={"old": True}, calculation_valid=True)
    app.invalidate_current_results(state)
    assert state.current_results is None
    assert not app.current_results_are_exportable(state)
    assert state.latest_results == {"old": True}


def test_large_span_default_secondary_section_reports_actionable_context() -> None:
    params = _large_21x36_params()
    with pytest.raises(ValueError) as exc_info:
        app.calculate_project_results(params)
    message = str(exc_info.value)
    assert "次梁计算失败" in message
    assert "弯矩过大" in message
    assert "次梁各跨=9,9,9,9 m" in message
    assert "截面 b×h=150×400 mm" in message
    assert "h0=360 mm" in message
    assert "增大次梁高度" in message


def test_manual_live_span_default_secondary_section_reports_actionable_context() -> None:
    params = _manual_20x28_params()
    with pytest.raises(ValueError) as exc_info:
        app.calculate_project_results(params)
    message = str(exc_info.value)
    assert "次梁计算失败" in message
    assert "次梁各跨=7,7,7,7 m" in message
    assert "单跨兼容计算跨度=7 m" in message
    assert "不要沿用默认小截面" in message


def test_larger_regular_span_sections_can_complete_after_section_adjustment() -> None:
    params = _large_21x36_params(
        slab_h_mm=120.0,
        slab_h0_mm=95.0,
        secondary_b_mm=250.0,
        secondary_h_mm=650.0,
        secondary_h0_mm=600.0,
        main_b_mm=350.0,
        main_h_mm=800.0,
        main_h0_mm=750.0,
    )
    results = app.calculate_project_results(params)
    tables = app.build_result_tables(results, params)
    assert results["matrix"].main.model.boundaries_m[-1] == pytest.approx(21.0)
    assert "计算数据总表" in build_calculation_book_tables(params, tables, results)


def test_unsupported_irregular_main_spans_are_still_rejected_with_layout_context() -> None:
    params = _params_with(main_spans_text="6,7,5", main_supports_text="pin,pin,pin,pin")
    with pytest.raises(ValueError) as exc_info:
        app.calculate_project_results(params)
    message = str(exc_info.value)
    assert "矩阵分析计算失败" in message
    assert "仅支持规则等跨主梁楼盖" in message
    assert "主梁=6,7,5 m" in message
    assert "支承数量等于跨数+1" in message


def test_extremely_narrow_beam_has_no_satisfactory_bar_layout() -> None:
    options = recommend_longitudinal_rebar(200.0, b_mm=50.0)
    assert options and not any(option.is_ok for option in options)
    assert all("不可布置" in option.evaluation or option.area_mm2 < 200 for option in options)


def test_positive_beam_moment_uses_t_section_and_negative_uses_rectangle() -> None:
    model = ContinuousMemberModel("次梁", (5.0, 5.0), EI, (10.0, 10.0), (5.0, 5.0), (0.3, 0.3, 0.3), ("pin",) * 3)
    envelope = analyze_member_envelope(model)
    design = design_control_sections(envelope, "beam", 250, 500, 450, 11.9, 300, 270, flange_width_mm=1000, flange_thickness_mm=100)
    assert design.loc[design["设计方向"] == "正弯矩", "截面类型判断"].str.contains("T形截面").all()
    assert design.loc[design["设计方向"] == "负弯矩", "截面类型判断"].str.contains("矩形截面").all()
    assert design["抗剪公式来源"].str.contains("需人工复核").all()


def test_high_moment_reports_compression_zone_limit_failure() -> None:
    model = ContinuousMemberModel("次梁", (4.0,), EI, (450.0,), (0.0,), (0.2, 0.2), ("pin", "pin"))
    envelope = analyze_member_envelope(model)
    design = design_control_sections(envelope, "beam", 200, 400, 350, 9.6, 300, 210, flange_width_mm=200, flange_thickness_mm=80)
    assert (design["是否满足"] == "不满足").any()
    assert design["需人工复核事项"].str.contains("ξb|受压区|截面不足").any()


def test_duplicate_element_id_is_rejected() -> None:
    with pytest.raises(ValueError, match="element_id"):
        solve_continuous_beam(
            [BeamNode(0, 0, True), BeamNode(1, 2), BeamNode(2, 4, True)],
            [BeamElement(1, 0, 1, EI), BeamElement(1, 1, 2, EI)],
        )


def test_point_load_at_support_is_treated_as_support_node_load() -> None:
    model = ContinuousMemberModel(
        "主梁", (4.0, 4.0), EI, (0.0, 0.0), (0.0, 0.0), (0.2, 0.2, 0.2), ("pin",) * 3,
        (MemberPointLoad(4.0, 0.0, 10.0, "支座节点荷载"),),
    )
    left = analyze_member(model, {0}, include_dead=False)
    right = analyze_member(model, {1}, include_dead=False)
    assert sum(abs(value) for value in left.load_vector) > 0
    assert sum(abs(value) for value in right.load_vector) > 0


def test_excel_sheet_names_remain_unique_after_truncation() -> None:
    tables = {
        "这是一个非常长且截断后相同的工作表名称_A": pd.DataFrame({"A": [1]}),
        "这是一个非常长且截断后相同的工作表名称_B": pd.DataFrame({"A": [2]}),
    }
    data = build_excel_workbook(tables)
    assert zipfile.is_zipfile(BytesIO(data))
    workbook = load_workbook(BytesIO(data), read_only=True)
    assert len(workbook.sheetnames) == len(set(workbook.sheetnames)) == 2


def test_word_wide_table_is_split_into_readable_subtables() -> None:
    data = build_word_report("宽表回归", {"宽表": pd.DataFrame([{f"列{i}": i for i in range(22)}])})
    with zipfile.ZipFile(BytesIO(data)) as archive:
        xml = archive.read("word/document.xml").decode("utf-8")
    assert "续表（同一数据表按列拆分）" in xml
    assert xml.count("<w:gridCol") >= 22


def test_calculation_book_exports_summary_tables_instead_of_full_matrix_dump() -> None:
    params = default_parameters().to_dict()
    results = app.calculate_project_results(params)
    tables = app.build_result_tables(results, params)
    report_tables = build_calculation_book_tables(params, tables, results)

    assert list(report_tables["计算数据总表"].columns) == REPORT_SUMMARY_COLUMNS
    assert {"计算数据总表", "复核说明", "主梁补充数据"} == set(report_tables)
    assert "主梁控制剪力" in report_tables["主梁补充数据"]["项目"].to_string()

    markdown = build_markdown_report("计算书回归", report_tables)
    assert "计算数据总表" in markdown
    assert "一级类别" in markdown
    assert "板总刚度矩阵摘要" not in markdown
    assert "节点位移表" not in markdown

    word = build_word_report("计算书回归", report_tables)
    with zipfile.ZipFile(BytesIO(word)) as archive:
        xml = archive.read("word/document.xml").decode("utf-8")
    assert "计算数据总表" in xml
    assert "板总刚度矩阵摘要" not in xml

    pdf = build_pdf_report({"project_name": "计算书回归", "date": "2026-06-30"}, params, tables, results)
    assert pdf.startswith(b"%PDF")


def test_run_scripts_use_an_available_port_instead_of_fixed_8501() -> None:
    root = Path(__file__).resolve().parents[1]
    for script_name in ("run_mac.command", "run_win.bat"):
        script = (root / script_name).read_text(encoding="utf-8")
        assert "--server.port 0" in script
        assert "http://localhost:8501" not in script
